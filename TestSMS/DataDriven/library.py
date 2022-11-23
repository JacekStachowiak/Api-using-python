import requests
import json
import jsonpath
import os
import openpyxl

class Common:
    
    def __init__(self, FileNamePath, SheetName):
        global wk
        global sk
        wk = openpyxl.load_workbook(FileNamePath, SheetName)
        sk = wk[SheetName]
        
    def row_count(self):
        rows = sk.max_row
        return rows

    def column_count(self):
        col = sk.max_column
        return col
    
    def key_names(self):
        c = sk.max_column
        li = []
        for i in range(1, c+1):
            cell = sk.cell(row=1, column=i)
            li.insert(i, cell.value)
        return li    
    
    def update_request_with_data(self, rowNumber, jsonRequest,keyList):
        c = sk.max_column
        for i in range(1, c+1):
            cell = sk.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value
        return jsonRequest            
            
        