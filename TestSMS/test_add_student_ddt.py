import requests
import json
import jsonpath
import os
import openpyxl

def test_add_students_ddt():
    url = 'https://thetestingworldapi.com/api/studentsDetails'
    f = open(os.path.abspath('add_student_ddt.json'), 'r')
    json_request = json.loads(f.read())
    
    wk = openpyxl.load_workbook(os.path.abspath('test_data.xlsx'))
    sk = wk['Arkusz1']
    rows = sk.max_row
 
    for i in range(2, rows+1): # 1 to nagłówek
        cell_first_name = sk.cell(row=i, column=1)
        cell_middle_name = sk.cell(row=i, column=2)
        cell_last_name = sk.cell(row=i, column=3)
        cell_date_of_birth = sk.cell(row=i, column=4)
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(url, json_request)
        print(response.status_code)
        print(response.text)
        assert response.status_code == 201
        
    
    
    